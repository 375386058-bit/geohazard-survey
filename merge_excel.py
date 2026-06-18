# -*- coding: utf-8 -*-
"""
地灾调查表 Excel 合并工具

用法:
  1. 把所有导出的 Excel 文件放在同一个文件夹
  2. 运行: python merge_excel.py
  3. 生成合并后的 地质灾害调查表_合并.xlsx

或者指定文件/文件夹:
  python merge_excel.py 文件1.xlsx 文件2.xlsx ...
  python merge_excel.py 文件夹路径

也可以合并从手机导出的 JSON 文件:
  python merge_excel.py --json *.json
"""

import sys
import os
import glob
import json
import re
from io import BytesIO

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


def parse_single_excel(filepath):
    """解析单个 Excel 文件，返回调查点列表"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    
    sites = []
    current_site = None
    current_section = None  # (seq, name)
    current_sub_rows = []
    expecting_data = False
    
    for row in rows:
        if not row:
            continue
        seq, name, content = (row[0] if len(row) > 0 else None,
                              row[1] if len(row) > 1 else None,
                              row[2] if len(row) > 2 else None)
        
        # Detect site header
        if name and isinstance(name, str) and name.startswith('调查点：'):
            if current_site:
                current_site['raw_rows'] = current_sub_rows
                sites.append(current_site)
            current_site = {
                'name': name.replace('调查点：', ''),
                'rows': [],
                'raw_rows': []
            }
            current_sub_rows = []
            expecting_data = False
            continue
        
        if not current_site:
            continue
        
        # Update time row
        if name and isinstance(name, str) and name.startswith('更新时间：'):
            continue
        
        current_sub_rows.append(list(row))
    
    if current_site:
        current_site['raw_rows'] = current_sub_rows
        sites.append(current_site)
    
    return sites


def merge_excel_files(filepaths, output_path):
    """合并多个 Excel 文件为一个"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '地质灾害调查表'
    
    # Header
    ws.append(['序号', '项目名称', '选项内容'])
    
    all_sites = []
    for fp in filepaths:
        try:
            sites = parse_single_excel(fp)
            for s in sites:
                # 去重
                if s['name'] not in [x['name'] for x in all_sites]:
                    all_sites.append(s)
            print(f"  [OK] {os.path.basename(fp)} -> {len(sites)} 个调查点")
        except Exception as e:
            print(f"  [FAIL] {os.path.basename(fp)} 解析失败: {e}")
    
    print(f"\n共合并 {len(all_sites)} 个调查点（已自动去重）\n")
    
    for site in all_sites:
        # Write site separator
        ws.append(['', f"调查点：{site['name']}", ''])
        for row in site['raw_rows']:
            ws.append(row)
        ws.append(['', '', ''])  # blank separator
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 60
    
    wb.save(output_path)
    print(f"[OK] 合并完成 -> {output_path}")
    return len(all_sites)


def json_to_excel_rows(site_data):
    """将 JSON 数据转为 Excel 行"""
    d = site_data.get('data', site_data)
    name = site_data.get('name', '未命名')
    updated = site_data.get('updatedAt', '')
    if updated:
        from datetime import datetime
        try:
            updated = datetime.fromtimestamp(updated / 1000).strftime('%Y/%m/%d %H:%M:%S')
        except:
            updated = str(updated)
    
    rows = [
        ['', f'调查点：{name}', ''],
        ['', f'更新时间：{updated}', ''],
        [1, '乡镇与野外编号', d.get('villageCode', '')],
        [2, '确定类型：', '□崩塌 □滑坡（灾害点） □孕灾点 □解译点'],
        ['', '', f'【已选：{"、".join(d.get("disasterType",[]))}】'] if d.get('disasterType') else None,
        [3, '总体地貌与微地貌', '总体地貌：□平原 □丘陵 □低山 □中山'],
        ['', '', f'  已选：{"、".join(d.get("landformMacro",[]))}'] if d.get('landformMacro') else None,
        ['', '', '微地貌：□陡崖 □陡坡 □缓坡 □平台'],
        ['', '', f'  已选：{"、".join(d.get("landformMicro",[]))}'] if d.get('landformMicro') else None,
        [4, '岩体结构', '□整体块状结构 □层状结构 □碎裂结构 □散体结构'],
        ['', '', f'已选：{"、".join(d.get("rockStructure",[]))}'] if d.get('rockStructure') else None,
        [5, '斜坡结构', '□土质 □岩质 □土石混合'],
        ['', '', f'已选：{"、".join(d.get("slopeStructure",[]))}'] if d.get('slopeStructure') else None,
        [6, '斜坡类型（石质斜坡填写）', '□顺向坡 □切向坡 □横向坡 □逆向坡 □近水平层斜坡 □块状岩体斜坡'],
        ['', '', f'已选：{"、".join(d.get("slopeType",[]))}'] if d.get('slopeType') else None,
        [7, '坡面形态', '□凹-凸型 □凸型 □阶梯型 □直型 □凹型'],
        ['', '', f'已选：{"、".join(d.get("slopeShape",[]))}'] if d.get('slopeShape') else None,
        [8, '岩层产状（倾向与倾角）', f'{d.get("stratumDip","____")}°<{d.get("stratumDipAngle","____")}°'],
        ['', '', f'长度：{d.get("stratumLength","____")} m'],
        ['', '', f'间距：{d.get("stratumSpacing","____")} m'],
        [9, '裂隙产状1（倾向与倾角）', f'{d.get("fissure1Dip","____")}°<{d.get("fissure1DipAngle","____")}°'],
        ['', '', f'长度：{d.get("fissure1Length","____")} m'],
        ['', '', f'间距：{d.get("fissure1Spacing","____")} m'],
        [10, '裂隙产状2（倾向与倾角）', f'{d.get("fissure2Dip","____")}°<{d.get("fissure2DipAngle","____")}°'],
        ['', '', f'长度：{d.get("fissure2Length","____")} m'],
        ['', '', f'间距：{d.get("fissure2Spacing","____")} m'],
        [11, '裂隙分割的岩石块度(长×宽×高)(m)',
         f'{d.get("rockBlockL","____")} × {d.get("rockBlockW","____")} × {d.get("rockBlockH","____")}'],
        [12, '岩体（土体）裂隙发育情况', f'裂隙组数：{d.get("fissureGroups","____")} 组'],
        ['', '', f'裂隙长 {d.get("fissureLength","____")} 米'],
        ['', '', f'宽 {d.get("fissureWidth","____")} 米'],
        ['', '', f'密度 {d.get("fissureDensity","____")} 条/米'],
        [13, '有无下滑的堆积物、如有记录其形状和体积', '形状：□扇形 □锥形 □其他'],
        ['', '', f'  已选：{"、".join(d.get("debrisShape",[]))}'] if d.get('debrisShape') else None,
        ['', '', f'体积：(长×宽×高)(m) {d.get("debrisVolumeL","____")} × {d.get("debrisVolumeW","____")} × {d.get("debrisVolumeH","____")}'],
        [14, '人类工程活动与土地利用类型',
         f'□切坡建房 □切坡修路 □村庄建设用地 □交通用地 □耕地 □林地{" □其他"+d["humanActivityOther"] if d.get("humanActivityOther") else ""}'],
        ['', '', f'已选：{"、".join(d.get("humanActivity",[]))}'] if d.get('humanActivity') else None,
        [15, '地表水',
         f'□河流 □水库{" □其他"+d["surfaceWaterOther"] if d.get("surfaceWaterOther") else ""}'],
        ['', '', f'已选：{"、".join(d.get("surfaceWater",[]))}'] if d.get('surfaceWater') else None,
        [16, '植被类型与覆盖率', '□乔木 □灌木 □草本'],
        ['', '', f'已选：{"、".join(d.get("vegetationType",[]))}'] if d.get('vegetationType') else None,
        ['', '', f'覆盖率：{d.get("vegetationCoverage","____")} %'],
        [17, '有无树木歪斜、建筑裂缝、地面裂缝', '□树木歪斜 □建筑裂缝 □地面裂缝'],
        ['', '', f'已选：{"、".join(d.get("damageSigns",[]))}'] if d.get('damageSigns') else None,
        [18, '威胁人数（人）与威胁财产（万元）', f'威胁人数：{d.get("threatPeople","____")} 人'],
        ['', '', f'威胁财产：{d.get("threatProperty","____")} 万元'],
        [19, '孕灾点类型', '□地质构造点 □工程地质岩组点 □崩滑易滑地层点 □其他点 □斜坡结构点'],
        ['', '', f'已选：{"、".join(d.get("disasterPointType",[]))}'] if d.get('disasterPointType') else None,
        ['', '变形破坏情况及其他补充内容：', d.get('extraNotes', '')],
        ['', '孕灾点沿途观测内容：', d.get('observationNotes', '')],
        ['', '', ''],
    ]
    return [r for r in rows if r is not None]


def merge_json_files(filepaths, output_path):
    """从 JSON 文件合并生成 Excel"""
    all_sites = {}
    
    for fp in filepaths:
        try:
            with open(fp, 'r', encoding='utf-8') as f:
                data = json.load(f)
            items = data if isinstance(data, list) else [data]
            for item in items:
                name = item.get('name', '未命名')
                if name not in all_sites:
                    all_sites[name] = item
                    print(f"  [OK] {os.path.basename(fp)} -> {name}")
                else:
                    print(f"  - {os.path.basename(fp)} -> {name}（重复，跳过）")
        except Exception as e:
            print(f"  [FAIL] {os.path.basename(fp)} 解析失败: {e}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '地质灾害调查表'
    ws.append(['序号', '项目名称', '选项内容'])
    
    for name, site in all_sites.items():
        rows = json_to_excel_rows(site)
        for row in rows:
            ws.append(row)
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 60
    
    wb.save(output_path)
    print(f"\n[OK] 合并完成 -> {output_path}")
    print(f"共 {len(all_sites)} 个调查点（已自动去重）")
    return len(all_sites)


def main():
    args = sys.argv[1:]
    
    use_json = False
    paths = []
    
    for a in args:
        if a == '--json':
            use_json = True
        elif '*' in a or '?' in a:
            paths.extend(glob.glob(a))
        else:
            paths.append(a)
    
    if not paths:
        # Default: look for Excel files in current directory
        cwd = os.getcwd()
        xlsx_files = glob.glob(os.path.join(cwd, '地质灾害调查表*.xlsx'))
        xlsx_files = [f for f in xlsx_files if '合并' not in f]
        
        if xlsx_files:
            paths = xlsx_files
            print(f"找到 {len(paths)} 个 Excel 文件:")
            for p in paths:
                print(f"  - {os.path.basename(p)}")
        else:
            # Try JSON
            json_files = glob.glob(os.path.join(cwd, '地灾调查数据*.json'))
            if json_files:
                paths = json_files
                use_json = True
                print(f"找到 {len(paths)} 个 JSON 文件:")
                for p in paths:
                    print(f"  - {os.path.basename(p)}")
            else:
                print("未找到任何 Excel 或 JSON 文件。")
                print("用法:")
                print("  python merge_excel.py 文件1.xlsx 文件2.xlsx")
                print("  python merge_excel.py --json 数据1.json 数据2.json")
                print("  python merge_excel.py --json *.json")
                sys.exit(1)
    
    if not paths:
        print("没有找到可合并的文件。")
        sys.exit(1)
    
    # Collect only valid files
    valid = []
    for p in paths:
        if os.path.isfile(p):
            valid.append(os.path.abspath(p))
        elif os.path.isdir(p):
            if use_json:
                valid.extend(glob.glob(os.path.join(p, '*.json')))
            else:
                valid.extend(glob.glob(os.path.join(p, '*.xlsx')))
    
    output_path = os.path.abspath('地质灾害调查表_合并.xlsx')
    
    print(f"\n开始合并 {len(valid)} 个文件...\n")
    
    if use_json:
        merge_json_files(valid, output_path)
    else:
        merge_excel_files(valid, output_path)


if __name__ == '__main__':
    main()
